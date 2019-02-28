# -*- coding: utf-8 -*-
"""
Created on Wed Feb 27 09:43:54 2019

@author: Mi6
"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import Canvas
import cv2,os,time
from PIL import ImageTk
import sqlite3
import dlib
import cognitive_face as CF
from global_variables import personGroupId 
window = tk.Tk()
#helv36 = tk.Font(family='Helvetica', size=36, weight='bold')
window.title("Face_Recogniser")
canvas = Canvas(width = 250, height = 250, bg = 'white')
canvas.pack(expand = YES, fill = BOTH)

image = ImageTk.PhotoImage(file = "E:/tws1.png")
canvas.create_image(350, 40, image = image, anchor = NW)

dialog_title = 'QUIT'
dialog_text = 'Are you sure?'

window.configure(background='blue')

window.grid_rowconfigure(0, weight=1)
window.grid_columnconfigure(0, weight=1)

message = tk.Label(window, text="Deep Learning enabled IP Camera Based Attendance System " ,bg="white"  ,fg="black"  ,width=50  ,height=1,font=('times', 18, 'italic bold ')) 

message.place(x=360, y=180)

lbl = tk.Label(window, text="Roll No",width=15  ,height=1  ,fg="black"  ,bg="orange" ,font=('times', 15, ' bold ') ) 
lbl.place(x=470, y=230)

txt = tk.Entry(window,width=25  ,bg="silver" ,fg="black",font=('times', 15, ' bold '),borderwidth=1,relief="solid")
txt.place(x=680, y=230)

lbl2 = tk.Label(window, text="Student Name",width=15  ,height=1  ,fg="black"  ,bg="orange" ,font=('times', 15, ' bold ')) 
lbl2.place(x=470, y=280)

txt2 = tk.Entry(window,width=25  ,bg="silver" ,fg="black",font=('times', 15, ' bold ') ,borderwidth=1,relief="solid" )
txt2.place(x=680, y=280)

lbl3 = tk.Label(window, text="Notification : ",width=15  ,height=1  ,fg="black"  ,bg="orange"  ,font=('times', 15, ' bold')) 
lbl3.place(x=470, y=330)

message = tk.Label(window, text="" ,bg="silver"  ,fg="black"  ,width=21  ,height=1 , activebackground = "yellow" ,font=('times', 15, ' bold '),borderwidth=1,relief="solid") 
message.place(x=680, y=330)

lbl4 = tk.Label(window, text="Enter the path",width=15  ,height=1  ,fg="black"  ,bg="orange" ,font=('times', 15, ' bold ')) 
lbl4.place(x=470, y=380)

txt3 = tk.Entry(window,width=25  ,bg="silver" ,fg="black",font=('times', 15, ' bold ') ,borderwidth=1,relief="solid" )
txt3.place(x=680, y=380)

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass
 
    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
 
    return False
def insertOrUpdate(Id, Name, roll) :                                            # this function is for database
    connect = sqlite3.connect("Face-DataBase")                                  # connecting to the database
    cmd = "SELECT * FROM Students WHERE ID = " + Id                             # selecting the row of an id into consideration
    cursor = connect.execute(cmd)
    isRecordExist = 0
    for row in cursor:                                                          # checking wheather the id exist or not
        isRecordExist = 1
    if isRecordExist == 1:                                                      # updating name and roll no
        connect.execute("UPDATE Students SET Name = ? WHERE ID = ?",(Name, Id))
        connect.execute("UPDATE Students SET Roll = ? WHERE ID = ?",(roll, Id))
    else:
    	params = (Id, Name, roll)                                               # insering a new student data
    	connect.execute("INSERT INTO Students(ID, Name, Roll) VALUES(?, ?, ?)", params)
    connect.commit()                                                            # commiting into the database
    connect.close()
def TakeImages():        
    roll=(txt.get())
    name=(txt2.get())
    if(name.isalpha()):
        cap = cv2.VideoCapture(0)
        detector = dlib.get_frontal_face_detector()
        Id = roll[-3:]
        insertOrUpdate(Id, name, roll)                                                  # calling the sqlite3 database
        folderName = "user" + Id                                                        # creating the person or user folder
        folderPath = os.path.join(os.path.dirname(os.path.realpath(__file__)), "dataset/"+folderName)
        if not os.path.exists(folderPath):
            os.makedirs(folderPath)

        sampleNum = 0
        while(True):
            ret, img = cap.read()                                                       # reading the camera input
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)                                # Converting to GrayScale
            dets = detector(img, 1)
            for i, d in enumerate(dets):                                                # loop will run for each face detected
                sampleNum += 1
                cv2.imwrite(folderPath + "/User." + Id + "." + str(sampleNum) + ".jpg",
                    img[d.top():d.bottom(), d.left():d.right()])                                                # Saving the faces
                cv2.rectangle(img, (d.left(), d.top())  ,(d.right(), d.bottom()),(0,255,0) ,2) # Forming the rectangle
                cv2.waitKey(200)                                                        # waiting time of 200 milisecond
            cv2.imshow('frame', img)                                                    # showing the video input from camera on window
            cv2.waitKey(1)
            if(sampleNum >= 20):                                                        # will take 20 faces
                break
        res="Image has been captured for id "+roll;
        cap.release()                                                                   # turning the webcam off
        cv2.destroyAllWindows()                                                         # Closing all the opened windows
        message.configure(text= res)
        Key = '7af26ce4f0db46a58f2280df501f27c3'
        CF.Key.set(Key)
        BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0'  # Replace with your regional Base URL
        CF.BaseUrl.set(BASE_URL)
        user="user"+Id;
        res = CF.person.create(personGroupId, str(user))
        extractId = str(user)[-3:]
        connect = sqlite3.connect("Face-DataBase")
        cmd = "SELECT * FROM Students WHERE ID = " + extractId
        cursor = connect.execute(cmd)
        isRecordExist = 0
        for row in cursor:                                                          # checking wheather the id exist or not
            isRecordExist = 1
        if isRecordExist == 1:                                                      # updating name and roll no
            connect.execute("UPDATE Students SET personID = ? WHERE ID = ?",(res['personId'], extractId))
    connect.commit()                                                            # commiting into the database
    connect.close()
    res="Person Id Created Sucessfully"
    message.configure(text= res)
  
def get_person_id():
	person_id = ''
	extractId = str(txt.get())[-3:]
	connect = sqlite3.connect("Face-DataBase")
	c = connect.cursor()
	cmd = "SELECT * FROM Students WHERE ID = " + extractId
	c.execute(cmd)
	row = c.fetchone()
	person_id = row[3]
	connect.close()
	return person_id    
def TrainImages():
    roll=(txt.get())
    Id = roll[-3:]
    user="user"+Id;
    Key = '7af26ce4f0db46a58f2280df501f27c3'
    CF.Key.set(Key)
    BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0'  # Replace with your regional Base URL
    CF.BaseUrl.set(BASE_URL)
    imageFolder ="E:/Autoattendance-Cognitive-master/dataset/" + user
    person_id = get_person_id()
    for filename in os.listdir(imageFolder):
        if filename.endswith(".jpg"):
        	print(filename)
        	imgurl =imageFolder+"/"+filename
        	res = CF.face.detect(imgurl)
        	if len(res) != 1:
        		print("No face detected in image")            
        	else:
        		res = CF.person.add_face(imgurl, personGroupId, person_id)
        		print(res)	
        	time.sleep(6)
    res = CF.person_group.train(personGroupId)
    print(res)
    if res.equals("{}"):
        message.configure(text="Image trained sucessfully")
    else:
        message.configure(text="Image not trained")
    

def getDateColumn():
    currentDate = time.strftime("%d_%m_%y")
    wb = load_workbook(filename = "reports.xlsx")
    sheet = wb['IT16']
    for i in range(1, len(sheet['A']) + 1):	
        col = get_column_letter(i)
        if sheet['%s%s'% (col,'1')].value == currentDate:
            return col
def TrackImages():
    conn = sqlite3.connect('Face-DataBase')
    c = conn.cursor()

#get current date
    currentDate = time.strftime("%d_%m_%y")

#create a workbook and add a worksheet
    if(os.path.exists('./reports.xlsx')):
        wb = load_workbook(filename = "reports.xlsx")
        sheet = wb['IT16']
    # sheet[ord() + '1']
        for col_index in range(1, 100):
            col = get_column_letter(col_index)
            if sheet['%s%s' % (col,1)].value is None:
                col2 = get_column_letter(col_index - 1)
                if sheet['%s%s' % (col2,1)].value != currentDate:
                    sheet['%s%s' % (col,1)] = currentDate
                break
            
        wb.save(filename = "reports.xlsx")
    else:
        wb = Workbook()
        dest_filename = 'reports.xlsx'
        c.execute("SELECT * FROM Students ORDER BY Roll ASC")
        ws1 = wb.active
        ws1.title = "IT16"
        ws1.append(('Roll Number', 'Name', currentDate))
        ws1.append(('', '', ''))
        while True:
            a = c.fetchone()
            if a == None:
                break
            else:
                ws1.append((a[2], a[1]))
        wb.save(filename = dest_filename)
    detector = dlib.get_frontal_face_detector()
    img = cv2.imread(str(txt3.get()))
    dets = detector(img, 1)
    if not os.path.exists('./Cropped_faces'):
      os.makedirs('./Cropped_faces')
    res=("detected = " + str(len(dets)))
    message.configure(text= res)
    for i, d in enumerate(dets):
      cv2.imwrite('./Cropped_faces/face' + str(i + 1) + '.jpg', img[d.top():d.bottom(), d.left():d.right()])
    currentDate = time.strftime("%d_%m_%y")
    wb = load_workbook(filename = "reports.xlsx")
    sheet = wb['IT16']
    Key = '7af26ce4f0db46a58f2280df501f27c3'
    CF.Key.set(Key)
    BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0'  # Replace with your regional Base URL
    CF.BaseUrl.set(BASE_URL)

    connect = connect = sqlite3.connect("Face-DataBase")
    c = connect.cursor()

    attend = [0 for i in range(600)]	


    directory = 'E:/Autoattendance-Cognitive-master/Cropped_faces'
    for filename in os.listdir(directory):
        if filename.endswith(".jpg"):
            imgurl = directory+"/"+ filename
            res = CF.face.detect(imgurl)
            if len(res) != 1:
                print ("No face detected.")	
                continue	
            faceIds = []
            for face in res:
                faceIds.append(face['faceId'])
            res = CF.face.identify(faceIds, personGroupId)
            print(filename)
            print( res)
            for face  in res:
                if not face['candidates']:
                    print("Unknown")
                else:
                   personId = face['candidates'][0]['personId']
                   c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
                   row = c.fetchone()
                   attend[int(row[0])] = 1
                   res=" "+( row[1] + " recognized")
                   message.configure(text= res)
        time.sleep(6)
    for row in range(3, len(sheet['A']) + 1):
        rn = sheet['A%s'% row].value
        if rn is not None:
            rn = rn[-3:]
            if attend[int(rn)] != 0:
                col = getDateColumn()
                sheet["%s%s" % (col, str(row))] = 1
                res="Attendance has been Marked Sucessfully"
                message.configure(text= res)
    wb.save(filename = "reports.xlsx")	
def viewAttendance():
    os.system("reports.xlsx")
                    	
  
takeImg = tk.Button(window, text="Capture Images", command=TakeImages  ,fg="black"  ,bg="orange"  ,width=20  ,height=2, activebackground = "Red" ,font=('times', 15, ' bold '))
takeImg.place(x=430, y=450)
trainImg = tk.Button(window, text="Train Images", command=TrainImages  ,fg="black"  ,bg="orange" ,width=20  ,height=2, activebackground = "Red" ,font=('times', 15, ' bold '))
trainImg.place(x=720, y=450)
trackImg = tk.Button(window, text="Mark Attendance", command=TrackImages  ,fg="black"  ,bg="orange"  ,width=20  ,height=2, activebackground = "Red" ,font=('times', 15, ' bold '))
trackImg.place(x=430, y=550)
viewatt=tk.Button(window, text="View Attendance", command=viewAttendance  ,fg="black"  ,bg="orange"  ,width=20  ,height=2, activebackground = "Red" ,font=('times', 15, ' bold '))
viewatt.place(x=720, y=550)
quitWindow = tk.Button(window, text="Quit", command=window.destroy  ,fg="black"  ,bg="orange"  ,width=20  ,height=2, activebackground = "Red" ,font=('times', 15, ' bold '))
quitWindow.place(x=580, y=640)
#copyWrite = tk.Text(window, background=window.cget("background"), borderwidth=0,font=('times', 30, 'italic bold underline'))
#copyWrite.tag_configure("superscript", offset=10)
#copyWrite.insert("insert", "Developed by Ashish","", "TEAM", "superscript")
#copyWrite.configure(state="disabled",fg="red"  )
#copyWrite.pack(side="left")
#copyWrite.place(x=800, y=750)
 
window.mainloop()